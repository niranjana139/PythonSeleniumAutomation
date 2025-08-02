import openpyxl

def readUsernameData():
    book = openpyxl.load_workbook("C:\\Users\\Netcom\\Desktop\\Niranjana Obsqura\\TestData.xlsx")
    sheet = book.active
    if "LoginPage" in book.sheetnames:
        sheet = book.worksheets[0]
        username = sheet.cell(row=1, column=1).value
    return username

def readPasswordData():
    book = openpyxl.load_workbook("C:\\Users\\Netcom\\Desktop\\Niranjana Obsqura\\TestData.xlsx")
    sheet = book.active
    if "LoginPage" in book.sheetnames:
        sheet = book.worksheets[0]
        password = sheet.cell(row=1, column=2).value
    return password



