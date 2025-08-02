import openpyxl


class FileOperations:

    def __init__(self):
        print("Initializing FileOperations")
        self.book = openpyxl.load_workbook("C:\\Users\\Netcom\\Desktop\\Niranjana Obsqura\\fileread.xlsx")
        self.sheet = self.book.active

    def readData(self):
        cell=self.sheet.cell(row=2, column=1)
        print(cell.value)


    def writeData(self):
        self.sheet.cell(row=6, column=1).value = "Hello World"
        self.sheet.cell(row=6, column=2).value = "hello123"
        print(self.sheet.cell(row=6, column=1).value," ",self.sheet.cell(row=6, column=2).value)

    def traverseRow(self):
        for i in range(1,self.sheet.max_row+1):
            for j in range(1,self.sheet.max_column+1):
                #Traverse through all i rows of j column
                print(self.sheet.cell(row=i, column=j).value)

    def performmisc(self):
        print("Value in B3 is ", self.sheet['B3'].value)
        print(self.sheet.max_row)  # returns number of rows
        print(self.sheet.max_column)  # returns number of columns

if __name__ == '__main__':
    files=FileOperations()
    #files.readData()
    #files.writeData()
    #files.readData()
    files.traverseRow()
    files.performmisc()