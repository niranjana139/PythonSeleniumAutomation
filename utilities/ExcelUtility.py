import openpyxl
from openpyxl.utils import get_column_letter


class ExcelUtility:
    def __init__(self, file_path):
        # Initialize with the file path to the Excel file
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)


    def get_string_data(self, row, col, sheet_name):
        """Get a string value from a specific cell (row, column)"""
        sheet = self.workbook[sheet_name]
        cell = sheet.cell(row=row, column=col)  # openpyxl uses 1-based indexing
        return str(cell.value)

    def get_integer_data(self, row, col, sheet_name):
        """Get an integer value from a specific cell (row, column)"""
        sheet = self.workbook[sheet_name]
        cell = sheet.cell(row=row, column=col)  # openpyxl uses 1-based indexing
        # Ensuring the cell contains a number and then casting to int
        return int(cell.value) if cell.value is not None else None

    def get_all_data(self, sheet_name):
        """Return all the data from a sheet as a 2D list"""
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data
